import openpyxl

work_book = openpyxl.load_workbook("products.xlsx")
sheet = work_book["Products"]

fx_range = work_book.defined_names["fx_rates"]
print(fx_range)
print(fx_range.destinations)

cells = []

for title, coord in fx_range.destinations:
    ws = work_book[title]
    cells.append(ws[coord])

print(cells)

max_row_str = str(sheet.max_row)
print(max_row_str)

for row in sheet["C3:C" + max_row_str]:
    for cell in row:
        cell.value = "=$B${0}*VLOOKUP($C$2, fx_rates, 2, False)".format(cell.row)
        cell.number_format = "#,##0,00"

for row in sheet["D3:D" + max_row_str]:
    for cell in row:
        cell.value = "=$B${0}*VLOOKUP($D$2, fx_rates, 2, False)".format(cell.row)
        cell.number_format = "#,##0,00"

for row in sheet["E3:E" + max_row_str]:
    for cell in row:
        cell.value = "=$B${0}*VLOOKUP($E$2, fx_rates, 2, False)".format(cell.row)
        cell.number_format = "#,##0,00"

# work_book.save("named_range.xlsx")

work_book.create_named_range("products", sheet, "$A$3:$B$" + max_row_str)
# work_book.save("named_range.xlsx")