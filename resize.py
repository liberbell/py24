import openpyxl
from openpyxl.styles import Alignment

work_book = openpyxl.Workbook()
sheet_obj = work_book.active

print(sheet_obj)
sheet_obj.title = "FirstSheet"
print(sheet_obj)

sheet_obj["C1"] = "A high row"
sheet_obj["D4"] = "A wide column"

sheet_obj.row_dimensions[1].height = 70
sheet_obj.column_dimensions["D"].width = 60

# work_book.save("dimensions.xlsx")

print(work_book.create_sheet(title="BrandNewSheet", index=0))

new_sheet = work_book.active
print(new_sheet)

# work_book.save("dimensions.xlsx")

new_sheet.merge_cells("A1:D3")
# work_book.save("dimensions.xlsx")

# new_sheet["C2"] = "Data in a merge cell"
new_sheet["A1"] = "Data in a merge cell"
# work_book.save("dimensions.xlsx")

new_sheet["A1"].alignment = Alignment(horizontal="center")
# work_book.save("dimensions.xlsx")

new_sheet["A1"].alignment = Alignment(vertical="center")
work_book.save("dimensions.xlsx")