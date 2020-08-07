import openpyxl
from openpyxl.formatting.rule import IconSetRule, DataBarRule 

# work_book = openpyxl.load_workbook("zomato_reviews.xlsx")
work_book = openpyxl.load_workbook("zomato-reviews.xlsx")

sheet = work_book.active

icon_set_rule = IconSetRule(icon_style="4Arrows", type="num", values=[1, 2, 3, 4])
print(sheet.max_row)

sheet.conditional_formatting.add("Q2:Q9558", icon_set_rule)
# work_book.save("zomato_iconset.xlsx")

work_book = openpyxl.load_workbook("zomato-reviews.xlsx")
sheet = work_book.active

data_bar_rule = DataBarRule(start_type="num", start_value=1, end_type="num", end_value=4, color="ff0000")

sheet.conditional_formatting.add("Q2:Q9558", data_bar_rule)
# work_book.save("zomato_databar.xlsx")