import openpyxl
from openpyxl.styles import PatternFill, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule

print(sheet.max_row)
print(sheet.max_column)

for row in sheet["L2:N101"]:
    for cell in row:
        cell.number_format = "#,##0"

# work_book.save("sales_basic_conditional.xlsx")

yellow_background = PatternFill(bgColor="00FFFF00")
diff_style = DifferentialStyle(fill = yellow_background)

rule = Rule(type="expression", dxf=diff_style)
rule.formula = ["$M1<70000"]

# rule1 = Rule(type="expression", dxf=diff_style)
# rule.formula = ["$M1<70000"]

print(sheet.calculate_dimension())

sheet.conditional_formatting.add(sheet.calculate_dimension(), rule)
# work_book.save("sales_basic_conditional.xlsx")

# color_scale_rule = ColorScaleRule(start_type="min", start_color="00FFFF00", end_type="max", end_color="00ff0000")

work_book = openpyxl.load_workbook("sales_record.xlsx")
sheet = work_book.active

for row in sheet["K2:N101"]:
    for cell in row:
        cell.number_format = "#,##0"

# sheet.conditional_formatting.add("M2:N101", color_scale_rule)
# work_book.save("sales_profit_colorscale.xlsx")

color_scale_rule = ColorScaleRule(start_type="percentile", start_value=0, start_color="F2B5EC",
                                  mid_type="percentile", mid_value=50, mid_color="FFFF66",
                                  end_type="percentile", end_value=90, end_color="81DC3B")

sheet.conditional_formatting.add("M2:N101", color_scale_rule)
# work_book.save("sales_profit_colorscale.xlsx")

