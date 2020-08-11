from openpyxl.utils import FORMULAE
import openpyxl

# print(FORMULAE)

work_book = openpyxl.Workbook()
sheet = work_book.active

sheet["A1"] = 21
sheet["A2"] = 11
sheet["A3"] = 7
sheet["A4"] = 9
sheet["A5"] = 6

sheet["C2"] = "SUM:"
sheet["D2"] = "=SUM(A1:A5)"

# work_book.save("formurae.xlsx")

sheet["C3"] = "PRODUCT:"
sheet["D3"] = "=PRODUCT(A1:A5)"
# work_book.save("formurae.xlsx")

sheet["C4"] = "COUNT:"
sheet["D4"] = "=COUNT(A1:A9)"

sheet["C5"] = "MEAN:"
sheet["D5"] = "=AVERAGE(A1:A5)"
# work_book.save("formurae.xlsx")

header = ["Cake", "Quantity", "Price", "Revenue"]

data = [["Chocolate", 18, 5],
        ["Cheesecake", 13, 4.5],
        ["Tres Leches", 16, 5.5],
        ["Carrot", 8, 4],
        ["Red Velvet", 9, 4.5]]

print(work_book.create_sheet("CakeSales", index=0))

cake_sales_sheet = work_book["CakeSales"]

cake_sales_sheet.append(header)

for row in data:
    cake_sales_sheet.append(row)

# work_book.save("formurae.xlsx")

max_row_str = str(cake_sales_sheet.max_row)
print(max_row_str)

for row in cake_sales_sheet["D2:D" + max_row_str]:
    for cell in row:
        cell.value = "=$B${0}*$C${0}".format(cell.row)

total_row_str = str(cake_sales_sheet.max_row + 2)
cake_sales_sheet["C" + total_row_str] = "Total Sales:"
cake_sales_sheet["D" + total_row_str] = "=SUM(D2:D" + max_row_str + ")"

# work_book.save("formurae.xlsx")

for row in cake_sales_sheet["C2:D" + max_row_str]:
    for cell in row:
        cell.number_format = "$#,##0.00"

cake_sales_sheet["D" + total_row_str].number_format = "$#,##0.00"
work_book.save("formurae.xlsx")