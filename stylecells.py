import openpyxl
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors
from openpyxl.styles import NamedStyle
from openpyxl.styles.fills import PatternFill

work_book = openpyxl.load_workbook("student_data.xlsx")
print(work_book.sheetnames)

sheet = work_book["Sheet"]
bold_font = Font(bold=True)
big_red_text = Font(color="FFFF0000", size=20)

center_aligned_text = Alignment(horizontal="center")
doubule_border_side = Side(border_style="double")

square_border = Border(top=doubule_border_side, right=doubule_border_side, bottom=doubule_border_side, left=doubule_border_side)

sheet["B2"].font = bold_font
sheet["B3"].font = big_red_text
sheet["C4"].alignment = center_aligned_text
sheet["C5"].border = square_border

# work_book.save(filename="styled.xlsx")

sheet["B7"].alignment = center_aligned_text
sheet["B7"].font = big_red_text
sheet["B7"].border = square_border
# work_book.save(filename="styled.xlsx")

custom_style = NamedStyle(name="header")
custom_style.font = Font(bold=True)

custom_style.border = Border(bottom=Side(border_style="thin"))
custom_style.alignment = Alignment(horizontal="center", vertical="center")

header_row = sheet[1]
for cell in header_row:
    cell.style = custom_style

# work_book.save(filename="styled.xlsx")

one_more_style = NamedStyle(name="highlight")
one_more_style.fill = PatternFill(fgColor=colors.Color("d7abcc"), patternType="lightHorizontal")

for cell in sheet["A"]:
    cell.style = one_more_style

# work_book.save(filename="styled.xlsx")