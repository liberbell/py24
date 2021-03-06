import openpyxl
import json


work_book = openpyxl.load_workbook("company_revenue.xlsx")
print(work_book.sheetnames)

sheet_obj = work_book["Revenue"]
print(sheet_obj.title)
print(sheet_obj["A1"])
print(sheet_obj["A1"].value)

cell = sheet_obj["B1"]
print(type(cell))
print(dir(cell))

print(cell.row)
print(cell.column)

print(cell.number_format)
print(cell.coordinate)
print(cell.data_type)

print(cell.value)

print(sheet_obj["A2"].value + ", based in " + sheet_obj["B2"].value + " has a revenue of $" + str (sheet_obj["C2"].value) + " billion.")
print(sheet_obj.cell(row=1, column=2))
print(sheet_obj.cell(row=1, column=2).value)

print(sheet_obj.max_row)
print(sheet_obj.max_column)

max_col = sheet_obj.max_column
for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row=1, column=i)
    print(cell_obj.value)

max_row = sheet_obj.max_row
for i in range(1, max_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    print(cell_obj.value)

print(sheet_obj["A1":"C2"])

for rows in sheet_obj["A1":"C2"]:
    for cell in rows:
        print(cell.coordinate, cell.value)
    print("---------------")

for row in sheet_obj.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3):
    print(row)

for value in sheet_obj.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3, values_only=True):
    print(value)

for column in sheet_obj.iter_cols(min_row=1, max_row=3, min_col=1, max_col=3, values_only=True):
    print(column)

revenues = {}

for row in sheet_obj.iter_rows(min_row=2, max_row=4, min_col=1, max_col=3, values_only=True):
    rep = row[0]
    rev_details = {
        "Country": row[1],
        "Revenue": row[2],
    }

    revenues[rep] = rev_details

print(json.dumps(revenues, indent=4, sort_keys=True))